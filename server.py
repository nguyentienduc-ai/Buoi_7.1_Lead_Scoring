import http.server
import socketserver
import json
import os
import urllib.parse
import webbrowser
import threading
import time
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PORT = 9090
DIRECTORY = os.path.dirname(os.path.abspath(__file__))
LOCAL_JSON = os.path.join(DIRECTORY, "scored_leads.json")
EXPORT_PATH = os.path.join(DIRECTORY, "scored_leads_final.xlsx")

def get_leads_data():
    if not os.path.exists(LOCAL_JSON):
        # Nếu chưa có dữ liệu, chạy sync tự động
        run_scoring_sync()
        
    try:
        with open(LOCAL_JSON, "r", encoding="utf-8") as f:
            records = json.load(f)
    except Exception as e:
        records = []
        
    # Tính toán thông số thống kê
    total = len(records)
    vip = sum(1 for r in records if r.get("status") == "VIP")
    junk = sum(1 for r in records if r.get("status") == "Junk")
    neutral = sum(1 for r in records if r.get("status") == "Neutral")
    
    last_modified = os.path.getmtime(LOCAL_JSON) if os.path.exists(LOCAL_JSON) else time.time()
    
    return {
        "records": records,
        "stats": {
            "total": total,
            "vip": vip,
            "junk": junk,
            "neutral": neutral
        },
        "last_modified": last_modified
    }

def run_scoring_sync():
    """Chạy script python đồng bộ dữ liệu"""
    script_path = os.path.join(DIRECTORY, "scripts", "score_leads.py")
    result = subprocess.run(["python", script_path], capture_output=True, text=True)
    if result.returncode != 0:
        raise Exception(f"Lỗi chạy score_leads.py: {result.stderr}")
    return True

def generate_excel_export():
    """Tạo file Excel được định dạng chuyên nghiệp từ dữ liệu JSON hiện tại"""
    if not os.path.exists(LOCAL_JSON):
        return False
        
    with open(LOCAL_JSON, "r", encoding="utf-8") as f:
        records = json.load(f)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Scoring Reports"
    
    # Đảm bảo hiển thị đường lưới (grid lines)
    ws.views.sheetView[0].showGridLines = True
    
    # Định nghĩa phong cách
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Navy Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name=font_family, size=10)
    data_font_bold = Font(name=font_family, size=10, bold=True)
    
    # Định dạng các dòng trạng thái
    status_fills = {
        "VIP": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),    # Soft Green
        "Junk": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),   # Soft Red
        "Neutral": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Soft Gray
    }
    
    status_fonts = {
        "VIP": Font(name=font_family, size=10, bold=True, color="375623"),
        "Junk": Font(name=font_family, size=10, bold=True, color="C65911"),
        "Neutral": Font(name=font_family, size=10, bold=True, color="595959")
    }
    
    # Biên viền mỏng
    thin_border_side = Side(style='thin', color='D3D3D3')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    headers = [
        "Mã KH (ID)", 
        "Tên khách hàng", 
        "Số điện thoại", 
        "Nhu cầu mô tả", 
        "Điểm AI", 
        "Phân loại", 
        "Lý do chấm điểm", 
        "Ghi chú kiểm duyệt", 
        "Người duyệt"
    ]
    
    # Ghi Headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        
    ws.row_dimensions[1].height = 28
    
    # Ghi Dữ liệu
    for idx, r in enumerate(records, start=2):
        reasons_str = ", ".join(r.get("reasons", []))
        comment = r.get("comment", "")
        reviewer = "Human Editor" if r.get("manual_override", False) else "AI Auto Scorer"
        
        row_data = [
            r.get("id"),
            r.get("ten_khach"),
            r.get("sdt"),
            r.get("nhu_cau_mo_ta"),
            r.get("score"),
            r.get("status"),
            reasons_str,
            comment,
            reviewer
        ]
        
        ws.append(row_data)
        ws.row_dimensions[idx].height = 22
        
        # Thiết lập styles cho từng ô trong dòng
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            
            # Căn chỉnh mặc định
            if col_idx in [1, 5]: # ID, Score
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [3, 6, 9]: # SDT, Status, Reviewer
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else: # Tên, Mô tả, Lý do, Comment
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Đảm bảo Số điện thoại là định dạng TEXT để không bị mất số 0
            if col_idx == 3:
                cell.number_format = '@'
                
            # Bọc dòng mô tả nhu cầu
            if col_idx == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Thiết lập định dạng màu sắc cho Phân loại (VIP, Junk, Neutral)
            if col_idx == 6:
                status_val = r.get("status", "Neutral")
                cell.fill = status_fills.get(status_val, status_fills["Neutral"])
                cell.font = status_fonts.get(status_val, status_fonts["Neutral"])
                
            # Đánh dấu in đậm cho cột Điểm số
            if col_idx == 5:
                cell.font = data_font_bold

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        col_idx = col[0].column
        
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val))
            else:
                max_len = max(max_len, len(val))
                
        # Giới hạn độ rộng cụ thể cho một số cột dài
        if col_idx == 4: # Mô tả nhu cầu
            ws.column_dimensions[col_letter].width = 50
        elif col_idx == 7: # Lý do chấm điểm
            ws.column_dimensions[col_letter].width = 35
        elif col_idx == 8: # Ghi chú kiểm duyệt
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(EXPORT_PATH)
    return True

class LeadScoringHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Tắt logs để tránh tràn terminal
        pass

    def send_json_response(self, status_code, data):
        response_bytes = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", str(len(response_bytes)))
        self.end_headers()
        self.wfile.write(response_bytes)

    def do_OPTIONS(self):
        # Hỗ trợ CORS preflight
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path

        if path == "/api/data":
            try:
                data = get_leads_data()
                self.send_json_response(200, data)
            except Exception as e:
                self.send_json_response(500, {"error": f"Lỗi lấy dữ liệu: {str(e)}"})

        elif path == "/api/export":
            try:
                success = generate_excel_export()
                if not success or not os.path.exists(EXPORT_PATH):
                    self.send_json_response(500, {"error": "Không thể tạo file Excel xuất."})
                    return
                
                with open(EXPORT_PATH, "rb") as f:
                    file_data = f.read()

                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="scored_leads_final.xlsx"')
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Content-Length", str(len(file_data)))
                self.end_headers()
                self.wfile.write(file_data)
            except Exception as e:
                self.send_json_response(500, {"error": f"Lỗi xuất Excel: {str(e)}"})

        elif path in ["/", "/index.html", "/dashboard.html"]:
            html_path = os.path.join(DIRECTORY, "dashboard.html")
            if not os.path.exists(html_path):
                self.send_response(404)
                self.end_headers()
                self.wfile.write(b"HTML dashboard file not found. Please create dashboard.html first.")
                return

            with open(html_path, "r", encoding="utf-8") as f:
                content = f.read()

            content_bytes = content.encode('utf-8')
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(content_bytes)))
            self.end_headers()
            self.wfile.write(content_bytes)
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")

    def do_POST(self):
        parsed_url = urllib.parse.urlparse(self.path)
        path = parsed_url.path

        if path == "/api/update":
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                payload = json.loads(post_data.decode('utf-8'))
                
                lead_id = payload.get("id")
                new_score = payload.get("score")
                new_status = payload.get("status")
                new_comment = payload.get("comment", "")
                
                if lead_id is None or new_score is None or new_status is None:
                    self.send_json_response(400, {"error": "Thiếu các trường thông tin bắt buộc (id, score, status)."})
                    return
                
                # Tải dữ liệu hiện tại
                with open(LOCAL_JSON, "r", encoding="utf-8") as f:
                    records = json.load(f)
                    
                updated = False
                for r in records:
                    if r["id"] == int(lead_id):
                        r["score"] = int(new_score)
                        r["status"] = new_status
                        r["comment"] = new_comment
                        r["manual_override"] = True
                        updated = True
                        break
                        
                if not updated:
                    self.send_json_response(404, {"error": f"Không tìm thấy khách hàng với ID {lead_id}."})
                    return
                    
                # Ghi lại file JSON
                with open(LOCAL_JSON, "w", encoding="utf-8") as f:
                    json.dump(records, f, ensure_ascii=False, indent=2)
                    
                self.send_json_response(200, {"status": "success", "message": "Cập nhật thông tin thành công!"})
            except Exception as e:
                self.send_json_response(500, {"error": f"Lỗi cập nhật dữ liệu: {str(e)}"})

        elif path == "/api/sync":
            try:
                run_scoring_sync()
                # Phản hồi dữ liệu mới sau khi sync
                data = get_leads_data()
                self.send_json_response(200, {
                    "status": "success", 
                    "message": "Đồng bộ và chấm điểm dữ liệu từ Google Sheets thành công!",
                    "data": data
                })
            except Exception as e:
                self.send_json_response(500, {"error": f"Lỗi đồng bộ dữ liệu: {str(e)}"})
        else:
            self.send_response(404)
            self.end_headers()
            self.wfile.write(b"Not Found")

def start_server():
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), LeadScoringHandler) as httpd:
        print(f"Lead Scoring Server is running at http://localhost:{PORT}")
        httpd.serve_forever()

def open_browser():
    time.sleep(1.5)
    print("Opening browser automatically...")
    webbrowser.open(f"http://localhost:{PORT}")

if __name__ == "__main__":
    browser_thread = threading.Thread(target=open_browser)
    browser_thread.daemon = True
    browser_thread.start()
    start_server()
